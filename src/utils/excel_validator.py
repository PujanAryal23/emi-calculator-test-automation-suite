"""openpyxl wrapper for the EMI calculator's exported workbook."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from src.utils.logging_setup import get_logger

log = get_logger(__name__)


@dataclass(frozen=True)
class ExcelSummary:
    """Parsed view of the downloaded workbook: header metadata, the
    schedule's header-row index, and the post-header data rows."""

    sheet_names: list[str]
    primary_sheet: str
    all_cells: list[list]
    schedule_header_row: int
    headers: list[str]
    schedule_rows: list[list]


@dataclass(frozen=True)
class MonthlyEntry:
    """One parsed row from the workbook's monthly amortization schedule."""

    month_index: int          # 1-based ordinal (1..N)
    month_label: str          # e.g. "May-2026"
    principal: Decimal
    interest: Decimal
    total_payment: Decimal
    balance: Decimal
    loan_paid_pct: Decimal


def load(path: Path | str) -> ExcelSummary:
    """Open the workbook and slice it into summary metadata + schedule rows.
    Raises `FileNotFoundError` if the file is missing or empty."""
    path = Path(path)
    if not path.exists() or path.stat().st_size == 0:
        log.error("excel export missing or empty: %s", path)
        raise FileNotFoundError(f"Excel export missing or empty: {path}")

    log.info("loading excel: %s", path)
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    if not rows:
        raise ValueError(f"Workbook {path} has no rows on sheet {sheet.title}")

    header_idx = _find_schedule_header(rows)
    headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    schedule = [r for r in rows[header_idx + 1:] if any(c is not None for c in r)]
    log.debug(
        "excel parsed: sheet=%s rows=%d header_row=%d schedule_rows=%d",
        sheet.title, len(rows), header_idx, len(schedule),
    )

    return ExcelSummary(
        sheet_names=wb.sheetnames,
        primary_sheet=sheet.title,
        all_cells=rows,
        schedule_header_row=header_idx,
        headers=headers,
        schedule_rows=schedule,
    )


def find_summary_cell(summary: ExcelSummary, label: str) -> Decimal | None:
    """Return the numeric value in column B for the summary row whose
    label (column A) contains `label` (case-insensitive substring match).
    Returns `None` if the label is not found."""
    needle = label.lower().strip()
    for row in summary.all_cells[:summary.schedule_header_row]:
        if not row or row[0] is None:
            continue
        if needle in str(row[0]).lower():
            value = _to_decimal(row[1] if len(row) > 1 else None)
            log.info("Summary cell %r → %s", label, value)
            return value
    log.info("Summary cell %r not found", label)
    return None


def monthly_schedule(summary: ExcelSummary) -> list[MonthlyEntry]:
    """Return every monthly schedule row as a list of `MonthlyEntry`,
    preserving the workbook's chronological order."""
    entries = [
        MonthlyEntry(
            month_index=int(row[0]),
            month_label=str(row[1]),
            principal=_to_decimal(row[2]),
            interest=_to_decimal(row[3]),
            total_payment=_to_decimal(row[4]),
            balance=_to_decimal(row[5]),
            loan_paid_pct=_to_decimal(row[6]),
        )
        for row in summary.schedule_rows
        if row[0] is not None and row[1] is not None
    ]
    log.info("Parsed %d monthly schedule entries from workbook", len(entries))
    return entries


def _find_schedule_header(rows) -> int:
    for idx, row in enumerate(rows):
        as_text = " ".join(str(c).lower() for c in row if c is not None)
        if "principal" in as_text and "interest" in as_text and "balance" in as_text:
            return idx
    return 0


def _to_decimal(value) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    cleaned = str(value).replace(",", "").replace("₹", "").replace("%", "").strip()
    if not cleaned:
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")
